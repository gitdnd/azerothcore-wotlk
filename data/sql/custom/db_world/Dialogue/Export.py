import openpyxl
from pathlib import Path
import math
import JsonifyDialogue as JsD

PrintAll = []
Flags = []
CurrentConversation = 0
CurrentScope = 0


def SplitVariants(text):
    tags = []
    text.split('&&')
def ExcelSearch():
    for d in Path('./').rglob('*.xlsx'):
        excelFile = openpyxl.load_workbook(d)
        try:
            for sheet in excelFile.worksheets:
                rows = sheet.max_row
                columns = sheet.max_column
                for i in range(1, rows + 1):
                    id = (sheet.cell(row = i, column = 1).value)
                    if id is None:
                        continue
                    text = (sheet.cell(row = i, column = 2).value)
                    flagsNeeded = (sheet.cell(row = i, column = 3).value)
                    flagsGiven = (sheet.cell(row = i, column = 4).value)
                    nextId = (sheet.cell(row = i, column = 5).value)

                    if CurrentScope != 0:
                        realId = sheet.title + "_" + CurrentScope + '_' + id
                    else:
                        realId = sheet.title + '_' + id
                    flagsNeeded = flagsNeeded.split()
                    flagsGiven = flagsGiven.split()
                    nextId = nextId.split()

                    if id == "END":
                        CurrentScope = 0
                    elif CurrentScope != 0:
                    if id[1] == "I":
                        CurrentBatch = 0
                        CurrentScope = id
                    elif id[1] == "B":
                        CurrentBatch = id
                    elif id[1] == "C":
                        CurrentScope = id
                    elif id[1] == "D":

                    elif id[1] == "T":
                        JsD.RefineGossip(realId, text, flagsNeeded, flagsGiven, nextId)

                    if text is not None:
                        PrintAll.append(sheet.title + "_" + id + '", "' + text)
                        if flagsNeeded is not None:
                            fN = flagsNeeded.split()
                            for f in fN:
                                Flags.append((sheet.title + "_" + f))
                        if flagsGiven is not None:
                            fG = flagsGiven.split()
                            for f in fG:
                                Flags.append((sheet.title + "_" + f))

        except:
            pass

ExcelSearch()

file = open('../elk_dialogue_text.sql', 'w')
file.write('''
DELIMITER ;;
DROP PROCEDURE IF EXISTS create_elk_dialogue_text;
CREATE PROCEDURE create_elk_dialogue_text ()
BEGIN
    DECLARE CONTINUE HANDLER FOR 1050 BEGIN END;
	CREATE TABLE elk_dialogue_text (
    id int NOT NULL AUTO_INCREMENT,
    text_name varchar(64) NOT NULL,
    dialogue TEXT NOT NULL,
	UNIQUE (id),
	UNIQUE (text_name)
   );
END;;
CALL create_elk_dialogue_text();;
DROP PROCEDURE  create_elk_dialogue_text;;

TRUNCATE TABLE elk_dialogue_text;
DROP PROCEDURE IF EXISTS add_elk_dialogue_text;
SET @dialogueAdded = 1;


CREATE PROCEDURE add_elk_dialogue_text (IN text_nameP VARCHAR(64), dialogueP TEXT)
BEGIN
   DECLARE CONTINUE HANDLER FOR 1050 BEGIN END;
	REPLACE INTO `elk_dialogue_text` (`id`, `text_name`, `dialogue`) VALUES
	(@dialogueAdded, text_nameP, dialogueP);
	REPLACE INTO `npc_text` (`ID`, `text0_0`, `Probability0`) VALUES
	(@dialogueAdded + 1000000, dialogueP, 1);
	SET @dialogueAdded = @dialogueAdded + 1;
END;;




CALL add_elk_dialogue_text ("DOTDOTDOT", "...");


''')
for item in PrintAll:
    file.write('CALL add_elk_dialogue_text ("'+item+'");\n')

file.close()


file = open('../elk_flags.sql', 'w')
file.write('''
DELIMITER ;;
DROP PROCEDURE IF EXISTS create_elk_flag;
CREATE PROCEDURE create_elk_flag ()
BEGIN
    DECLARE CONTINUE HANDLER FOR 1050 BEGIN END;
	CREATE TABLE elk_flag (
    id int NOT NULL AUTO_INCREMENT,
    flag_name varchar(64) NOT NULL,
	UNIQUE (id)
   );
END;;
CALL create_elk_flag();;
DROP PROCEDURE  create_elk_flag;;


TRUNCATE TABLE elk_flag;
DROP PROCEDURE IF EXISTS add_elk_flag;
SET @flagAdded = 1;


CREATE PROCEDURE add_elk_flag (IN flag_nameP VARCHAR(64))
BEGIN
   DECLARE CONTINUE HANDLER FOR 1050 BEGIN END;
	REPLACE INTO `elk_dialogue_text` (`id`, `flag_name`) VALUES
	(@flagAdded, flag_nameP);
	SET @dialogueAdded = @dialogueAdded + 1;
END;;

CALL add_elk_flag ("TEST_FLAG");

''')
for item in Flags:
    file.write('CALL add_elk_flag ("'+item+'");\n')

file.close()
